import os
import django
import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "test.settings")
django.setup()

from main.models import FAQ, Lesson, Class, Unit

wb = openpyxl.load_workbook("faq_list.xlsx")
ws = wb.worksheets[0]

# 講座名を登録
l = []
for i in range(2, ws.max_row):
    if ws[f"E{i}"].value == None:
        break
    if ws[f"E{i}"].value not in l:
        l.append(ws[f"E{i}"].value)

l = list(map(lambda x: Class(name=x), l))
Class.objects.bulk_create(l)

# 単元名を登録
l = []
for i in range(2, ws.max_row):
    if ws[f"E{i}"].value == None:
        break
    pair = (ws[f"E{i}"].value, ws[f"F{i}"].value)
    if pair not in l:
        l.append(pair)

l = list(map(lambda x: Unit(name=x[1], parent=Class.objects.get(name=x[0])), l))
Unit.objects.bulk_create(l)

# 学習項目名を登録
l = []
for i in range(2, ws.max_row):
    if ws[f"E{i}"].value == None:
        break
    pair = (ws[f"F{i}"].value, ws[f"G{i}"].value)
    if pair not in l:
        l.append(pair)

l = list(map(lambda x: Lesson(name=x[1], parent=Unit.objects.get(name=x[0])), l))
Lesson.objects.bulk_create(l)
    
# 質問を登録
l = []
for i in range(2, ws.max_row):
    if ws[f"E{i}"].value == None:
        break
    triple = (ws[f"G{i}"].value, ws[f"I{i}"].value, ws[f"J{i}"].value)
    if triple not in l:
        l.append(triple)

l = list(map(lambda x: FAQ(question=x[1], answer=x[2], lesson=Lesson.objects.get(name=x[0])), l))
FAQ.objects.bulk_create(l)
